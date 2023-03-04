import os
import re
import numpy as np

import openpyxl as ox
from openpyxl.utils import get_column_letter
import pyexcel as p

# from openpyxl import load_workbook


directory = 'Admin/Downloads'
search_text = 'ОЛ-11'
print('Ищем: ', search_text)

for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        filepath = os.path.join(directory, filename)
        wb = ox.load_workbook(filename=filepath, read_only=True)
        ws = wb[wb.sheetnames[0]]  # первый лист
        row_max = ws.max_row
        column_max = ws.max_column
        print('In file: ', filename, '\n Columns:', column_max, '\n Rows: ', row_max)
        # s66_value = ws['S66'].value
        # print(f'{filename}: {s66_value}')
        row_min = 1
        column_min = 1

        arr_A = []
        arr_indexes_A = []
        for i in range(row_max + 1):
            word_cell_A = 'A' + str(int(i))
            arr_indexes_A.append(word_cell_A)
            # print(word_cell_A)
            data_from_cell_A = ws[word_cell_A].value
            arr_A.append(data_from_cell_A)

        NoneType = type(None)
        index_to_remove = []
        for j in range(len(arr_A)):
            if type(arr_A[j]) == NoneType:
                index_to_remove.append(j)


        arr_A = np.delete(arr_A, index_to_remove)
        arr_indexes_A = np.delete(arr_indexes_A, index_to_remove)


        #Очистим массивы индексов и данных из ячеек от мусора
        arrays_n = []
        next_clean_arr = []
        for i in range(len(arr_A)):
            nums = '123456'
            if arr_A[i] not in nums:
                next_clean_arr.append(i)
                # if int(arr_A[i]) > int(arr_A[i-1]):
                #     arrays_n.append(arr_A[i])
                # else:
                #     break
        arr_A = np.delete(arr_A, next_clean_arr)
        arr_indexes_A = np.delete(arr_indexes_A, next_clean_arr)

        #Функция создания массивов индексов и данных из ячеек
        

        while column_min <= column_max:
            row_min_min = row_min
            row_max_max = row_max
            while row_min_min <= row_max_max:
                row_min_min = str(row_min_min)

                word_column = get_column_letter(column_min)
                word_column = str(word_column)
                word_cell = word_column + row_min_min
                word_cell_next = word_column + str(int(row_min_min) + 1)

                data_from_cell = ws[word_cell].value
                data_from_cell = str(data_from_cell)
                # print(data_from_cell)
                regular = search_text
                result = re.findall(regular, data_from_cell)
                if len(result) > 0:
                    print('Нашли в ячейке: ', word_cell)
                    data_from_cell_next = ws[word_cell_next].value
                    print(data_from_cell_next)
                    print(arr_A)
                    print(arr_indexes_A)
                    print(arrays_n)
                    print(next_clean_arr)

                row_min_min = int(row_min_min)
                row_min_min += 1
            column_min += 1